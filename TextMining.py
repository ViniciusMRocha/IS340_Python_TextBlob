from openpyxl import *
from textblob import TextBlob
from statistics import *
from wordcloud import WordCloud
import matplotlib.pyplot as plt

exbook = load_workbook('restaurants.xlsx')
sheet1 = exbook.active

londonRatings = []
londonTextList = []
london5 = []
london4 = []
london3 = []
london2 = []
london1 = []

otherRatings = []
otherTextList = []
other5 = []
other4 = []
other3 = []
other2 = []
other1 = []

wordList = []
goodList =[]
badList = []

#       ----------CLEANING UP LONDON----------

for i in range (2, 20000):
    location=sheet1.cell(column=10,row=i).value
    if location is None:
        location='None'
    elif 'Lond' in location:
        sheet1.cell(column=10,row=i).value='LONDON'
    elif 'lond' in location:
        sheet1.cell(column=10,row=i).value='LONDON'
    elif 'LOND' in location:
        sheet1.cell(column=10,row=i).value='LONDON'


# ----------SEPARATE LONDON AND OTHER LOCATIONS----------

for i in range(2, 20000):
    location = sheet1.cell(column=10, row=i).value
    if location is None:
        location = "None"

    elif 'LONDON' in location:
        londonRating = float(sheet1.cell(column=11, row=i).value)
        londonRatings.append(londonRating)
        if londonRating == 5:
            london5.append(londonRating)
            good = sheet1.cell(column=8, row=i).value
            goodList.append(good)
        elif londonRating == 4:
            london4.append(londonRating)
            good = sheet1.cell(column=8, row=i).value
            goodList.append(good)
        elif londonRating == 3:
            london3.append(londonRating)
        elif londonRating == 2:
            london2.append(londonRating)
            bad = sheet1.cell(column=8, row=i).value
            badList.append(bad)
        elif londonRating == 1:
            london1.append(londonRating)
            bad = sheet1.cell(column=8, row=i).value
            badList.append(bad)
        londonReview = TextBlob(sheet1.cell(column=8, row=i).value)
        londonTextList.append(londonReview.sentiment.polarity)

    elif 'LONDON' not in location:
        otherRating = float(sheet1.cell(column=11, row=i).value)
        otherRatings.append(otherRating)
        if otherRating ==5:
            other5.append(otherRating)
            good = sheet1.cell(column=8, row=i).value
            goodList.append(good)
        elif otherRating == 4:
            other4.append(otherRating)
            good = sheet1.cell(column=8, row=i).value
            goodList.append(good)
        elif otherRating == 3:
            other3.append(otherRating)
        elif otherRating == 2:
            other2.append(otherRating)
            bad = sheet1.cell(column=8, row=i).value
            badList.append(bad)
        elif otherRating == 1:
            other1.append(otherRating)
            bad = sheet1.cell(column=8, row=i).value
            badList.append(bad)
        otherReview = TextBlob(sheet1.cell(column=8, row=i).value)
        otherTextList.append(otherReview.sentiment.polarity)

#       ----------STATISTICS LONDON STAR----------

print("-"*16, "LONDONER'S STAR RATINGS", "-"*16)
print("")
print("Number of 5 Star Ratings: ", len(london5))
print("Number of 4 Star Ratings: ", len(london4))
print("Number of 3 Star Ratings: ", len(london3))
print("Number of 2 Star Ratings: ", len(london2))
print("Number of 1 Star Ratings: ", len(london1))
print("_________________________________")
print("Total Number of Ratings:  ", len(londonRatings))
print("")
print("")
londonRatingMax = max(londonRatings)
londonRatingMin = min(londonRatings)
londonRatingAvg = round(mean(londonRatings), 4)
londonRatingMode = mode(londonRatings)
londonRatingSD = round(stdev(londonRatings), 4)
print("""
Maximum Londoner Rating:                           %1.4f
Minimum Londoner Rating:                           %1.4f
Average Londoner Rating:                           %1.4f
Londoner Rating Mode:                              %1.4f
Londoner Rating Standard Deviation:                %1.4f""" %(londonRatingMax, londonRatingMin, londonRatingAvg, londonRatingMode, londonRatingSD))
print("")
print("")
print("")

#      ----------STATISTICS LONDON SENTIMENT----------

print("-"*14, "LONDONER'S SENTIMENT ANALYSIS", "-"*13)
print("")
londonSentMax = round(max(londonTextList), 4)
londonSentMin = round(min(londonTextList), 4)
londonSentAvg = round(mean(londonTextList), 4)
londonSentMode = mode(londonTextList)
londonSentSD = round(stdev(londonTextList), 4)
print("""
Maximum Londoner Sentiment Polarity:              %1.4f
Minimum Londoner Sentiment Polarity:              %1.4f
Average Londoner Sentiment Polarity:              %1.4f
Londoner Sentiment Polarity Mode:                 %1.4f
Londoner Sentiment Polarity Standard Deviation:   %1.4f """ %(londonSentMax, londonSentMin, londonSentAvg, londonSentMode, londonSentSD))
print("")
print("")
print("")


#        ----------STATISTICS OTHER STAR----------

print("-"*17, "TRAVELER STAR RATINGS", "-"*17)
print("")
print("Number of 5 Star Ratings: ", len(other5))
print("Number of 4 Star Ratings: ", len(other4))
print("Number of 3 Star Ratings: ", len(other3))
print("Number of 2 Star Ratings: ", len(other2))
print("Number of 1 Star Ratings: ", len(other1))
print("_________________________________")
print("Total Number of Ratings:  ", len(otherRatings))
print("")
print("")
otherRatingMax = max(otherRatings)
otherRatingMin = min(otherRatings)
otherRatingAvg = round(mean(otherRatings), 4)
otherRatingMode = mode(otherRatings)
otherRatingSD = round(stdev(otherRatings), 4)
print("""
Maximum Traveler Rating:                           %1.4f
Minimum Traveler Rating:                           %1.4f
Average Traveler Rating                            %1.4f
Traveler Rating Mode:                              %1.4f 
Traveler Rating Standard Deviation:                %1.4f""" %(otherRatingMax, otherRatingMin, otherRatingAvg, otherRatingMode, otherRatingSD))
print("")
print("")
print("")

#         ----------STATISTICS OTHER SENTIMENT----------

print("-"*15, "TRAVELER SENTIMENT ANALYSIS", "-"*15)
print("")
otherSentMax = round(max(otherTextList), 4)
otherSentMin = round(min(otherTextList), 4)
otherSentAvg = round(mean(otherTextList), 4)
otherSentMode = mode(otherTextList)
otherSentSD = round(stdev(otherTextList), 4)
print("""
Maximum Travelers Sentiment Polarity:              %1.4f
Minimum Travelers Sentiment Polarity:              %1.4f
Average Travelers Sentiment Polarity:              %1.4f
Travelers Sentiment Polarity Mode:                 %1.4f
Travelers Sentiment Polarity Standard Deviation:   %1.4f""" %(otherSentMax,otherSentMin,otherSentAvg,otherSentMode,otherSentSD))

wordcloud = WordCloud(width = 4000, height = 2000).generate(str(goodList))
plt.imshow(wordcloud)
plt.axis("off")
plt.show()

wordcloud = WordCloud(width = 4000, height = 2000).generate(str(badList))
plt.imshow(wordcloud)
plt.axis("off")
plt.show()
